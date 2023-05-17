import openpyxl as op
import pprint as pp
import json
import os


file_dir = "Данные для парсинга-20230514T121156Z-001"


def passport_parse(file_name):
    passport_dict = {}
    wb = op.load_workbook(file_name, data_only=True)
    sheet = wb.active

    max_rows = sheet.max_row


    for item in range(14, max_rows + 1):
        key = sheet.cell(row=item, column=2).value
        if key != None:
            result = sheet.cell(row=item, column=3).value
            if result == None:
                result = "-"
                passport_dict[key] = result
            else:
                passport_dict[key] = result

    pp.pprint(passport_dict)

    with open("passport_json.json", "w", encoding="utf-8") as file:
        for key, result in passport_dict.items():
            json.dump(
                passport_dict,
                file,
                sort_keys=False,
                indent=4,
                separators=(",", ": "),
                ensure_ascii=False,
            )


for item in os.listdir(file_dir):
    f = os.path.join(file_dir, item)
    passport_parse(f)
